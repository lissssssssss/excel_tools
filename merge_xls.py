# -*- coding: utf-8 -*-
"""
按月合并 .xls 文件 (含图片 + 单元格格式 + 合并单元格) 为单个 .xlsx 文件。

工作流:
  1. 扫描源目录, 按文件名前缀 (如 "3.", "4.") 分组。
  2. 对每个 .xls 文件:
     - 用 xlrd (formatting_info=True) 读取单元格、字体、对齐、行高、列宽、合并单元格;
     - 直接解析 .xls 的 OLE/BIFF/OfficeArt 二进制结构, 抽取每张图片
       (JPEG/PNG/DIB/TIFF) 以及每张图片所在的 (行, 列);
  3. 把所有数据行 (按文件顺序) 顺序追加到一个新的 .xlsx:
     - 字体、对齐 (含自动换行)、数字格式、边框、行高、列宽 都尽量保留;
     - 合并单元格按源文件→输出文件的行号映射重新合并;
     - 图片嵌入到对应单元格 (PNG 保留为 PNG, 不强制转 JPEG);
     - 末尾再追加两列 "源文件" / "源文件行号" 方便排查。

注意:
  - 输入是 .xls (Excel 97-2003)。输出是 .xlsx (纯 Python 内嵌图片只能输出 .xlsx)。
  - 不需要安装 LibreOffice / Office, 纯 Python。

用法:
  python merge_xls.py [源目录] [输出目录]
默认值:
  源目录   : /Users/lishengsheng/Documents/铝合金代发
  输出目录 : 源目录 + "/_合并"
"""
from __future__ import annotations

import argparse
import io
import logging
import re
import struct
import sys
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

import olefile
import xlrd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


# ====== 日志配置 ============================================================

def setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("merge_xls")
    logger.setLevel(logging.DEBUG)
    logger.handlers.clear()

    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)-7s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(logging.INFO)
    sh.setFormatter(fmt)
    logger.addHandler(sh)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger


# ====== .xls BIFF / OfficeArt 解析 (用于图片抽取) ===========================
# 参考: MS-XLS, MS-ODRAW

_OART_CONTAINERS = {0xF000, 0xF001, 0xF002, 0xF003, 0xF004, 0xF005}
_BLIP_TYPES = {
    0xF01A: "emf",
    0xF01B: "wmf",
    0xF01C: "pict",
    0xF01D: "jpeg",
    0xF01E: "png",
    0xF01F: "dib",
    0xF029: "tiff",
    0xF02A: "jpeg",
}
# 栅格图: 当 recInstance 为下列值时, payload 头部是 32 字节 (两个 UID), 否则 16 字节; 之后还有 1 字节 tag
_TWO_UID_INSTANCES = {0x6E1, 0x46B, 0x6E3, 0x7A9, 0x6E5}


def _iter_biff(stream: bytes):
    i, n = 0, len(stream)
    last_type = None
    last_data = bytearray()
    while i + 4 <= n:
        rec_type, rec_len = struct.unpack_from("<HH", stream, i)
        i += 4
        if i + rec_len > n:
            break
        chunk = stream[i : i + rec_len]
        i += rec_len
        if rec_type == 0x003C and last_type is not None:
            last_data.extend(chunk)
            continue
        if last_type is not None:
            yield last_type, bytes(last_data)
        last_type = rec_type
        last_data = bytearray(chunk)
    if last_type is not None:
        yield last_type, bytes(last_data)


def _decode_blip_payload(rec_type: int, rec_inst: int, payload: bytes):
    fmt = _BLIP_TYPES[rec_type]
    if rec_type in (0xF01A, 0xF01B, 0xF01C):
        return None, fmt
    head_len = 33 if rec_inst in _TWO_UID_INSTANCES else 17
    if len(payload) < head_len:
        return None, fmt
    return payload[head_len:], fmt


def _walk_oart_for_blips(data: bytes, out: list):
    i, n = 0, len(data)
    while i + 8 <= n:
        ver_inst, rt, rl = struct.unpack_from("<HHI", data, i)
        inst = (ver_inst >> 4) & 0xFFF
        i += 8
        if i + rl > n:
            break
        payload = data[i : i + rl]
        i += rl
        if rt in _BLIP_TYPES:
            out.append((rt, inst, payload))
        elif rt in _OART_CONTAINERS:
            _walk_oart_for_blips(payload, out)
        elif rt == 0xF007:  # OfficeArtFBSE - BLIP 嵌在头 36 字节后面
            if len(payload) >= 36:
                _walk_oart_for_blips(payload[36:], out)


def _parse_anchor(payload: bytes):
    """解析 OfficeArtClientAnchorSheet。
    格式 (18 字节): flag c1 dx1 r1 dy1 c2 dx2 r2 dy2 (各 2 字节, 小端 unsigned)
      - r1/c1, r2/c2: 图片左上/右下"参考格子"行列号 (0-based)
      - dy1/dy2: 在 r1/r2 内的 Y 偏移, 单位 = 1/256 行高
      - dx1/dx2: 在 c1/c2 内的 X 偏移, 单位 = 1/1024 列宽
    返回 (center_y, center_x, r1, c1, r2, c2):
      - center_y, center_x 为图片几何中心 (浮点的"格子坐标")
      - 归属格子 = (floor(center_y), floor(center_x))
    """
    if len(payload) < 18:
        return None
    _, c1, dx1, r1, dy1, c2, dx2, r2, dy2 = struct.unpack_from("<HHHHHHHHH", payload, 0)
    top_y = r1 + dy1 / 256.0
    bot_y = r2 + dy2 / 256.0
    left_x = c1 + dx1 / 1024.0
    right_x = c2 + dx2 / 1024.0
    return (
        (top_y + bot_y) / 2.0,
        (left_x + right_x) / 2.0,
        r1, c1, r2, c2,
    )


def _parse_fopt_pib(num_props: int, payload: bytes):
    head = payload[: num_props * 6]
    for k in range(num_props):
        opid_full = struct.unpack_from("<H", head, k * 6)[0]
        val = struct.unpack_from("<I", head, k * 6 + 2)[0]
        opid = opid_full & 0x3FFF
        if opid == 0x0104:
            return val
    return None


def _walk_oart_for_shapes(data: bytes, out: list):
    i, n = 0, len(data)
    while i + 8 <= n:
        ver_inst, rt, rl = struct.unpack_from("<HHI", data, i)
        inst = (ver_inst >> 4) & 0xFFF
        i += 8
        if i + rl > n:
            break
        payload = data[i : i + rl]
        i += rl
        if rt == 0xF004:  # OfficeArtSpContainer
            anchor = None
            pib = None
            j, m = 0, len(payload)
            while j + 8 <= m:
                v2, rt2, rl2 = struct.unpack_from("<HHI", payload, j)
                inst2 = (v2 >> 4) & 0xFFF
                j += 8
                if j + rl2 > m:
                    break
                p2 = payload[j : j + rl2]
                j += rl2
                if rt2 == 0xF010:
                    anchor = _parse_anchor(p2)
                elif rt2 == 0xF00B:
                    pib = _parse_fopt_pib(inst2, p2)
            if anchor is not None and pib is not None:
                out.append((anchor, pib))
        elif rt in _OART_CONTAINERS:
            _walk_oart_for_shapes(payload, out)


def extract_images_per_sheet(xls_path: Path, logger: logging.Logger):
    """从 .xls 中按 sheet 抽取图片。
    返回 dict: {sheet_idx: {row0: (fmt, image_bytes)}}, 以及总诊断信息。
    sheet_idx 与 xlrd `book.sheet_by_index(idx)` 一致。
    """
    diagnostics = {"blip_total": 0, "shape_total_by_sheet": {}}
    try:
        ole = olefile.OleFileIO(str(xls_path))
    except Exception as e:
        logger.error("[%s] 不是合法的 OLE 复合文档: %s", xls_path.name, e)
        return {}, diagnostics
    try:
        wb_stream_name = None
        for s in ole.listdir():
            if s and s[-1].lower() in ("workbook", "book"):
                wb_stream_name = "/".join(s)
                break
        if wb_stream_name is None:
            logger.error("[%s] 未找到 Workbook 流", xls_path.name)
            return {}, diagnostics
        workbook_bytes = ole.openstream(wb_stream_name).read()
    finally:
        ole.close()

    # 解析 BIFF 流, 用 BOF (0x0809) / EOF (0x000A) 划分每个 substream:
    #   BOF.substream_type:
    #     0x0005 = workbook globals (这里有 MsoDrawingGroup)
    #     0x0010 = worksheet
    #     0x0020 = chart
    #     ...
    # MsoDrawingGroup (0xEB) 在 globals 流; MsoDrawing (0xEC) 在每个 sheet 流。
    # sheet_idx 从 0 开始计数 (和 xlrd 的 sheet 顺序一致)。
    mso_dg_group = bytearray()
    mso_drawings_per_sheet: dict = {}
    current_sheet_idx = -1  # -1 = globals 或未进入任何 sheet
    next_sheet_idx = 0
    for rt, rd in _iter_biff(workbook_bytes):
        if rt == 0x0809 and len(rd) >= 4:
            substream_type = struct.unpack_from("<H", rd, 2)[0]
            if substream_type == 0x0005:
                current_sheet_idx = -1
            else:
                current_sheet_idx = next_sheet_idx
                next_sheet_idx += 1
        elif rt == 0x000A:  # EOF
            current_sheet_idx = -1
        elif rt == 0x00EB:
            mso_dg_group.extend(rd)
        elif rt == 0x00EC and current_sheet_idx >= 0:
            mso_drawings_per_sheet.setdefault(current_sheet_idx, bytearray()).extend(rd)

    blips_raw: list = []
    _walk_oart_for_blips(bytes(mso_dg_group), blips_raw)
    diagnostics["blip_total"] = len(blips_raw)
    blip_table: list = []
    for rt, inst, payload in blips_raw:
        img_bytes, fmt = _decode_blip_payload(rt, inst, payload)
        blip_table.append((fmt, img_bytes))

    result_per_sheet: dict = {}
    for sheet_idx, drawing_bytes in mso_drawings_per_sheet.items():
        shapes: list = []
        _walk_oart_for_shapes(bytes(drawing_bytes), shapes)
        diagnostics["shape_total_by_sheet"][sheet_idx] = len(shapes)
        result_per_sheet[sheet_idx] = _assign_shapes_to_rows(
            shapes, blip_table, xls_path, sheet_idx, logger,
        )

    return result_per_sheet, diagnostics


def _assign_shapes_to_rows(
    shapes: list, blip_table: list, xls_path: Path, sheet_idx: int,
    logger: logging.Logger,
):
    """把同一 sheet 的形状转成 {row0: (fmt, image_bytes)}。
    用图片几何中心定位归属行; 同行多张时, 多余的图片顺移到下面 3 行内的空行 (人工
    粘贴位置不准的常见情形); 都被占用则丢弃并记日志。"""

    pending: list = []
    skipped_vector = 0
    skipped_oob = 0
    for anchor, pib in shapes:
        if pib < 1 or pib > len(blip_table):
            skipped_oob += 1
            continue
        fmt, img_bytes = blip_table[pib - 1]
        if img_bytes is None:
            skipped_vector += 1
            continue
        center_y, _center_x, r1, _c1, r2, _c2 = anchor
        target_row = int(center_y)
        if target_row < 0:
            target_row = 0
        pending.append((target_row, fmt, img_bytes, center_y, r1, r2))

    pending.sort(key=lambda x: (x[0], x[3]))

    result: dict = {}
    overflow: list = []
    for target_row, fmt, img_bytes, cy, _r1, _r2 in pending:
        if target_row in result:
            overflow.append((target_row, fmt, img_bytes, cy))
        else:
            result[target_row] = (fmt, img_bytes)

    _OVERFLOW_WINDOW = 3
    shifts: list = []
    dropped: list = []
    for target_row, fmt, img_bytes, cy in overflow:
        actual = None
        for delta in range(1, _OVERFLOW_WINDOW + 1):
            candidate = target_row + delta
            if candidate not in result:
                actual = candidate
                break
        if actual is not None:
            result[actual] = (fmt, img_bytes)
            shifts.append((target_row, actual))
        else:
            dropped.append(target_row)

    tag = f"{xls_path.name} sheet#{sheet_idx}"
    if skipped_vector:
        logger.warning("[%s] 跳过 %d 个矢量图片 (emf/wmf/pict)", tag, skipped_vector)
    if skipped_oob:
        logger.warning("[%s] 跳过 %d 个 pib 越界的形状", tag, skipped_oob)
    if shifts:
        from collections import Counter
        by_origin = Counter(s[0] for s in shifts)
        for orig_r, count in sorted(by_origin.items()):
            actual_rows = ",".join(str(t + 1) for o, t in shifts if o == orig_r)
            logger.warning(
                "[%s] 第 %d 行多出 %d 张图片, 顺移到 Excel 行 %s",
                tag, orig_r + 1, count, actual_rows,
            )
    if dropped:
        from collections import Counter
        by_origin = Counter(dropped)
        for orig_r, count in sorted(by_origin.items()):
            logger.warning(
                "[%s] 第 %d 行有 %d 张多余图片, 后面 %d 行均已占用, 已丢弃",
                tag, orig_r + 1, count, _OVERFLOW_WINDOW,
            )
    logger.debug(
        "[%s] 形状=%d, 已定位行数=%d, 顺移=%d, 丢弃=%d",
        tag, len(shapes), len(result), len(shifts), len(dropped),
    )
    return result


# ====== xlrd → openpyxl 格式转换 ============================================

# xlrd 水平对齐
_HOR_ALIGN = {
    0: None, 1: "left", 2: "center", 3: "right", 4: "fill",
    5: "justify", 6: "centerContinuous", 7: "distributed",
}
# xlrd 垂直对齐
_VER_ALIGN = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}

# xlrd 边框样式 → openpyxl 名称
_BORDER_STYLE = {
    0: None, 1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
    5: "thick", 6: "double", 7: "hair", 8: "mediumDashed",
    9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot",
    12: "mediumDashDotDot", 13: "slantDashDot",
}

# Excel 默认调色板 (足够覆盖 .xls 用到的颜色)
_DEFAULT_PALETTE = {
    0: (0, 0, 0), 1: (255, 255, 255), 2: (255, 0, 0), 3: (0, 255, 0),
    4: (0, 0, 255), 5: (255, 255, 0), 6: (255, 0, 255), 7: (0, 255, 255),
    8: (0, 0, 0), 9: (255, 255, 255), 10: (255, 0, 0), 11: (0, 255, 0),
    12: (0, 0, 255), 13: (255, 255, 0), 14: (255, 0, 255), 15: (0, 255, 255),
    16: (128, 0, 0), 17: (0, 128, 0), 18: (0, 0, 128), 19: (128, 128, 0),
    20: (128, 0, 128), 21: (0, 128, 128), 22: (192, 192, 192), 23: (128, 128, 128),
    24: (153, 153, 255), 25: (153, 51, 102), 26: (255, 255, 204), 27: (204, 255, 255),
    28: (102, 0, 102), 29: (255, 128, 128), 30: (0, 102, 204), 31: (204, 204, 255),
    32: (0, 0, 128), 33: (255, 0, 255), 34: (255, 255, 0), 35: (0, 255, 255),
    36: (128, 0, 128), 37: (128, 0, 0), 38: (0, 128, 128), 39: (0, 0, 255),
    40: (0, 204, 255), 41: (204, 255, 255), 42: (204, 255, 204), 43: (255, 255, 153),
    44: (153, 204, 255), 45: (255, 153, 204), 46: (204, 153, 255), 47: (255, 204, 153),
    48: (51, 102, 255), 49: (51, 204, 204), 50: (153, 204, 0), 51: (255, 204, 0),
    52: (255, 153, 0), 53: (255, 102, 0), 54: (102, 102, 153), 55: (150, 150, 150),
    56: (0, 51, 102), 57: (51, 153, 102), 58: (0, 51, 0), 59: (51, 51, 0),
    60: (153, 51, 0), 61: (153, 51, 102), 62: (51, 51, 153), 63: (51, 51, 51),
}


def _xls_color_to_hex(book, colour_index):
    """xlrd 调色板索引 → openpyxl 接受的 'RRGGBB' 字符串; 无色返回 None。"""
    if colour_index is None or colour_index in (0x7FFF, 64, 65):
        return None
    rgb = book.colour_map.get(colour_index) if hasattr(book, "colour_map") else None
    if rgb is None:
        rgb = _DEFAULT_PALETTE.get(colour_index)
    if rgb is None:
        return None
    return "{:02X}{:02X}{:02X}".format(*rgb)


def _xls_cell_to_python(cell, datemode):
    ct = cell.ctype
    v = cell.value
    if ct in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if ct == xlrd.XL_CELL_TEXT:
        return v
    if ct == xlrd.XL_CELL_NUMBER:
        if v == int(v):
            return int(v)
        return v
    if ct == xlrd.XL_CELL_DATE:
        try:
            tup = xlrd.xldate_as_tuple(v, datemode)
        except Exception:
            return v
        import datetime
        if tup[:3] == (0, 0, 0):
            return datetime.time(*tup[3:])
        if tup[3:] == (0, 0, 0):
            return datetime.date(*tup[:3])
        return datetime.datetime(*tup)
    if ct == xlrd.XL_CELL_BOOLEAN:
        return bool(v)
    if ct == xlrd.XL_CELL_ERROR:
        return f"#ERR{v}"
    return v


class StyleCache:
    """缓存 (book, xf_index) → (Font, Alignment, Border, NumberFormat)。"""

    def __init__(self):
        self._cache: dict = {}

    def get(self, book, xf_index: int):
        key = (id(book), xf_index)
        if key in self._cache:
            return self._cache[key]
        try:
            xf = book.xf_list[xf_index]
        except (IndexError, AttributeError):
            v = (None, None, None, "General")
            self._cache[key] = v
            return v

        font_obj = book.font_list[xf.font_index]
        font_kwargs = {}
        if font_obj.name:
            font_kwargs["name"] = font_obj.name
        if font_obj.height:
            font_kwargs["size"] = font_obj.height / 20.0
        if font_obj.bold:
            font_kwargs["bold"] = True
        if font_obj.italic:
            font_kwargs["italic"] = True
        if font_obj.underline_type:
            font_kwargs["underline"] = "single"
        color_hex = _xls_color_to_hex(book, font_obj.colour_index)
        if color_hex:
            font_kwargs["color"] = color_hex
        font = Font(**font_kwargs) if font_kwargs else None

        a = xf.alignment
        align_kwargs = {}
        h = _HOR_ALIGN.get(a.hor_align)
        if h:
            align_kwargs["horizontal"] = h
        v = _VER_ALIGN.get(a.vert_align)
        if v:
            align_kwargs["vertical"] = v
        if a.text_wrapped:
            align_kwargs["wrap_text"] = True
        if a.indent_level:
            align_kwargs["indent"] = a.indent_level
        if a.rotation and a.rotation != 0xFF:
            align_kwargs["text_rotation"] = a.rotation
        if a.shrink_to_fit:
            align_kwargs["shrink_to_fit"] = True
        alignment = Alignment(**align_kwargs) if align_kwargs else None

        b = xf.border
        sides = {}
        for side_name, line_attr, color_attr in (
            ("left", "left_line_style", "left_colour_index"),
            ("right", "right_line_style", "right_colour_index"),
            ("top", "top_line_style", "top_colour_index"),
            ("bottom", "bottom_line_style", "bottom_colour_index"),
        ):
            ls = _BORDER_STYLE.get(getattr(b, line_attr))
            if ls:
                col = _xls_color_to_hex(book, getattr(b, color_attr))
                sides[side_name] = Side(style=ls, color=col)
        border = Border(**sides) if sides else None

        try:
            number_format = book.format_map[xf.format_key].format_str or "General"
        except (KeyError, AttributeError):
            number_format = "General"

        result = (font, alignment, border, number_format)
        self._cache[key] = result
        return result


def _apply_style(cell, font, alignment, border, number_format):
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format and number_format != "General":
        cell.number_format = number_format


# ====== 单文件解析 ==========================================================

@dataclass
class ParsedRow:
    cells: list                           # [(value, xf_index), ...]
    image: tuple | None                   # (fmt, image_bytes) 或 None
    src_file: str
    src_sheet: str
    src_row_0based: int                   # 在源 sheet 中的行号 (xlrd 0-based)
    src_row_1based: int                   # Excel 行号 (1-based)
    row_height_pt: float | None           # 行高 (磅), None 表示用默认


@dataclass
class ParsedFile:
    """实际是 "已解析的 sheet": 一个 .xls 文件可能产生多个 ParsedFile (每 sheet 一个)。"""
    src_file: str
    src_sheet: str
    sheet_idx: int
    src_path: Path
    book: xlrd.book.Book                  # workbook 共享 (XF 索引是 workbook 级)
    header_cells: list                    # [(value, xf_index)]
    rows: list[ParsedRow] = field(default_factory=list)
    image_col: int = -1
    tracking_col: int = -1                # 快递单号列
    n_data_rows_total: int = 0
    n_blank_rows_skipped: int = 0
    col_widths_chars: dict = field(default_factory=dict)
    merges: list = field(default_factory=list)


def _detect_image_col(header_values: list) -> int:
    for idx, h in enumerate(header_values):
        if isinstance(h, str) and "图" in h:
            return idx
    return -1


def _detect_tracking_col(header_values: list) -> int:
    """识别 '快递单号' 列。优先精确匹配 '快递单号', 否则模糊匹配含 '快递' 或 '单号'。"""
    fuzzy = -1
    for idx, h in enumerate(header_values):
        if not isinstance(h, str):
            continue
        s = h.strip()
        if "快递单号" in s:
            return idx
        if fuzzy < 0 and ("快递" in s or "单号" in s or "运单" in s):
            fuzzy = idx
    return fuzzy


def _last_data_row(sh, max_cols: int) -> int:
    """返回最后一个含非空值的行 (0-based, 含); 没有数据则返回 0 (仅表头)。"""
    last = 0
    cols = min(max_cols, sh.ncols)
    for r in range(sh.nrows - 1, -1, -1):
        for c in range(cols):
            v = sh.cell_value(r, c)
            if v not in (None, "", " "):
                return r
    return last


def _last_data_col(sh, header_values: list) -> int:
    """返回实际的最后一列 (0-based, 含)。
    formatting_info=True 会把空但有格式的列也算进 ncols, 这里用表头里
    最后一个非空值定位真实数据列。"""
    last_header_col = -1
    for c, v in enumerate(header_values):
        if v not in (None, "", " "):
            last_header_col = c
    return last_header_col


def parse_xls_file(path: Path, logger: logging.Logger) -> list[ParsedFile]:
    """解析一个 .xls 文件的所有 (可见) sheet, 每个 sheet 返回一个 ParsedFile。
    失败或无可用 sheet 返回 []。"""
    try:
        book = xlrd.open_workbook(str(path), formatting_info=True, on_demand=False)
    except Exception as e:
        logger.error("[%s] 无法用 xlrd 打开: %s", path.name, e)
        logger.debug(traceback.format_exc())
        return []

    if book.nsheets == 0:
        logger.error("[%s] 没有任何 sheet", path.name)
        return []

    # 整个 workbook 一次性把图片按 sheet 抽取出来
    images_by_sheet: dict = {}
    image_diag: dict = {}
    try:
        images_by_sheet, image_diag = extract_images_per_sheet(path, logger)
    except Exception as e:
        logger.error("[%s] 解析图片失败: %s", path.name, e)
        logger.debug(traceback.format_exc())

    blip_total = image_diag.get("blip_total", 0)
    shape_total_by_sheet = image_diag.get("shape_total_by_sheet", {})

    parsed_sheets: list[ParsedFile] = []
    for sheet_idx in range(book.nsheets):
        sh = book.sheet_by_index(sheet_idx)
        # 跳过隐藏 sheet (visibility: 0=可见, 1=隐藏, 2=极深隐藏)
        visibility = getattr(sh, "visibility", 0)
        if visibility != 0:
            logger.info(
                "[%s] sheet#%d '%s' 是隐藏的, 跳过", path.name, sheet_idx, sh.name,
            )
            continue
        tag = f"{path.name} sheet#{sheet_idx} '{sh.name}'"
        ps = _parse_one_sheet(
            path=path, book=book, sh=sh, sheet_idx=sheet_idx, tag=tag,
            sheet_images=images_by_sheet.get(sheet_idx, {}),
            n_shapes=shape_total_by_sheet.get(sheet_idx, 0),
            blip_total=blip_total,
            logger=logger,
        )
        if ps is not None:
            parsed_sheets.append(ps)

    if not parsed_sheets:
        logger.warning("[%s] 没有可用的 sheet", path.name)
    return parsed_sheets


def _parse_one_sheet(
    path: Path, book, sh, sheet_idx: int, tag: str,
    sheet_images: dict, n_shapes: int, blip_total: int,
    logger: logging.Logger,
) -> ParsedFile | None:
    if sh.nrows == 0:
        logger.warning("[%s] 是空的, 跳过", tag)
        return None

    raw_header = [_xls_cell_to_python(sh.cell(0, c), book.datemode) for c in range(sh.ncols)]
    last_col = _last_data_col(sh, raw_header)
    if last_col < 0:
        logger.warning("[%s] 表头全为空, 跳过", tag)
        return None
    header_values = raw_header[: last_col + 1]
    image_col = _detect_image_col(header_values)
    tracking_col = _detect_tracking_col(header_values)
    if image_col < 0:
        logger.warning("[%s] 表头里没找到 '图' 字, 不嵌入图片. 表头=%s", tag, header_values)
    if tracking_col < 0:
        logger.warning("[%s] 表头里没找到 '快递单号' 列, 该 sheet 不参与去重", tag)

    last_row = _last_data_row(sh, last_col + 1)
    if last_row < 1:
        logger.warning("[%s] 没有数据行 (除了表头), 跳过", tag)
        return None

    n_cols = last_col + 1
    header_cells = [
        (header_values[c], sh.cell_xf_index(0, c) if hasattr(sh, "cell_xf_index") else 0)
        for c in range(n_cols)
    ]

    parsed = ParsedFile(
        src_file=path.name,
        src_sheet=sh.name,
        sheet_idx=sheet_idx,
        src_path=path,
        book=book,
        header_cells=header_cells,
        image_col=image_col,
        tracking_col=tracking_col,
    )

    for c, ci in (sh.colinfo_map or {}).items():
        if c < n_cols and ci.width and not ci.hidden:
            parsed.col_widths_chars[c] = ci.width / 256.0

    for rlo, rhi, clo, chi in (sh.merged_cells or []):
        if rhi - 1 <= last_row and chi - 1 < n_cols:
            parsed.merges.append((rlo, rhi, clo, chi))
        else:
            logger.debug(
                "[%s] 合并单元格 (%d,%d,%d,%d) 超出数据范围 (last_row=%d, n_cols=%d), 忽略",
                tag, rlo, rhi, clo, chi, last_row, n_cols,
            )

    images = dict(sheet_images)
    for r in list(images.keys()):
        if r > last_row:
            logger.warning(
                "[%s] 图片锚点在 Excel 第 %d 行, 已超出数据末尾 (Excel %d 行), 该图片被丢弃",
                tag, r + 1, last_row + 1,
            )
            images.pop(r, None)

    n_rows_with_image = 0
    for r in range(1, last_row + 1):
        cells = []
        for c in range(n_cols):
            v = _xls_cell_to_python(sh.cell(r, c), book.datemode)
            xfi = sh.cell_xf_index(r, c) if hasattr(sh, "cell_xf_index") else 0
            cells.append((v, xfi))

        non_image_vals = [
            v for c, (v, _x) in enumerate(cells)
            if c != image_col and v not in (None, "", " ")
        ]
        img = images.get(r) if image_col >= 0 else None

        if not non_image_vals and img is None:
            parsed.n_blank_rows_skipped += 1
            logger.debug("[%s] 第 %d 行空白, 跳过", tag, r + 1)
            continue
        if not non_image_vals and img is not None:
            logger.warning("[%s] 第 %d 行只有图片没有任何文本, 仍然保留", tag, r + 1)

        row_height = None
        ri = (sh.rowinfo_map or {}).get(r)
        if ri is not None and ri.height:
            row_height = ri.height / 20.0

        parsed.rows.append(
            ParsedRow(
                cells=cells,
                image=img,
                src_file=path.name,
                src_sheet=sh.name,
                src_row_0based=r,
                src_row_1based=r + 1,
                row_height_pt=row_height,
            )
        )
        if img is not None:
            n_rows_with_image += 1
        parsed.n_data_rows_total += 1

    logger.info(
        "[%s] 数据行=%d, 嵌入图片=%d, 跳过空白=%d  (BLIP仓库=%d, 形状=%d, 已定位=%d)",
        tag, parsed.n_data_rows_total, n_rows_with_image,
        parsed.n_blank_rows_skipped, blip_total, n_shapes, len(images),
    )
    return parsed


# ====== 写出 .xlsx ==========================================================

# 图片在输出中显示尺寸的盒子 (像素)
_IMG_BOX_PX = 100
_IMG_COL_WIDTH_CHARS = round(_IMG_BOX_PX / 7, 1)
_IMG_ROW_HEIGHT_PT = round(_IMG_BOX_PX / 1.333, 1)
# PIL 缩略图最大尺寸 (像素). 单元格显示 100x100, 这里给 2 倍清晰度便于放大。
_THUMBNAIL_BOX_PX = 220


def _prepare_image_bytes(image_bytes: bytes, fmt: str):
    """缩略到 _THUMBNAIL_BOX_PX. PNG 保留 PNG (含透明), 其余转 JPEG。"""
    try:
        with PILImage.open(io.BytesIO(image_bytes)) as im:
            mode = im.mode
            if mode == "P":
                im = im.convert("RGBA")
            elif mode == "CMYK":
                im = im.convert("RGB")
            im.thumbnail((_THUMBNAIL_BOX_PX, _THUMBNAIL_BOX_PX), PILImage.LANCZOS)
            buf = io.BytesIO()
            keep_alpha = im.mode in ("RGBA", "LA") or fmt == "png"
            if keep_alpha:
                if im.mode not in ("RGBA", "LA"):
                    im = im.convert("RGBA")
                im.save(buf, format="PNG", optimize=True)
                ext = "png"
            else:
                if im.mode != "RGB":
                    im = im.convert("RGB")
                im.save(buf, format="JPEG", quality=85, optimize=True)
                ext = "jpeg"
            buf.seek(0)
            return buf, ext
    except Exception:
        return io.BytesIO(image_bytes), fmt


def _normalize_tracking(value) -> str | None:
    """规范化快递单号: 去前后空格 + 内部空格、unicode 全角/半角统一、大小写。
    空值返回 None (即不参与去重)。"""
    if value is None:
        return None
    s = str(value)
    # 去除所有空白 (前后 + 内部)
    s = "".join(s.split())
    if not s:
        return None
    # 大小写统一
    return s.upper()


def write_merged_xlsx(
    out_path: Path,
    parsed_files: list[ParsedFile],
    logger: logging.Logger,
):
    if not parsed_files:
        logger.warning("没有任何输入文件可写入 %s", out_path.name)
        return

    style_cache = StyleCache()

    canonical = parsed_files[0]
    canonical_header = [v for v, _ in canonical.header_cells]
    canonical_image_col = canonical.image_col
    canonical_tracking_col = canonical.tracking_col
    n_canon_cols = len(canonical_header)

    # 同一文件下的 sheet 数量统计 (用于决定是否在 "源文件" 里附 sheet 名)
    file_sheet_count: dict = defaultdict(int)
    for pf in parsed_files:
        file_sheet_count[pf.src_file] += 1

    for pf in parsed_files[1:]:
        cur_header = [v for v, _ in pf.header_cells]
        if cur_header[: len(canonical_header)] != canonical_header[: len(cur_header)]:
            logger.warning(
                "[%s / %s] 表头与第一个 sheet 不一致, 仍按列序合并。\n  规范: %s\n  本表: %s",
                pf.src_file, pf.src_sheet, canonical_header, cur_header,
            )
        if pf.image_col != canonical_image_col and pf.image_col >= 0:
            logger.warning(
                "[%s / %s] 图片列序号 (%d) 与第一个 sheet (%d) 不同",
                pf.src_file, pf.src_sheet, pf.image_col, canonical_image_col,
            )
        if pf.tracking_col != canonical_tracking_col and pf.tracking_col >= 0:
            logger.warning(
                "[%s / %s] 快递单号列序号 (%d) 与第一个 sheet (%d) 不同",
                pf.src_file, pf.src_sheet, pf.tracking_col, canonical_tracking_col,
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    # ---- 表头 ----
    extended_header = list(canonical_header) + ["源文件", "源文件行号"]
    for col_idx, (val, xfi) in enumerate(canonical.header_cells, start=1):
        cell = ws.cell(row=1, column=col_idx, value=val)
        font, align, border, nfmt = style_cache.get(canonical.book, xfi)
        _apply_style(cell, font, align, border, nfmt)
        if cell.font is None or not cell.font.bold:
            base = cell.font or Font()
            cell.font = Font(
                name=base.name, size=base.size, bold=True,
                italic=base.italic, color=base.color, underline=base.underline,
            )
    extra1 = ws.cell(row=1, column=n_canon_cols + 1, value="源文件")
    extra2 = ws.cell(row=1, column=n_canon_cols + 2, value="源文件行号")
    bold_font = Font(bold=True)
    extra1.font = bold_font
    extra2.font = bold_font

    # ---- 列宽 ----
    for c in range(n_canon_cols):
        col_letter = get_column_letter(c + 1)
        if c == canonical_image_col:
            ws.column_dimensions[col_letter].width = max(
                _IMG_COL_WIDTH_CHARS,
                canonical.col_widths_chars.get(c, _IMG_COL_WIDTH_CHARS),
            )
        else:
            w = canonical.col_widths_chars.get(c)
            if w:
                ws.column_dimensions[col_letter].width = w
    ws.column_dimensions[get_column_letter(n_canon_cols + 1)].width = 36
    ws.column_dimensions[get_column_letter(n_canon_cols + 2)].width = 12

    # 表头行高
    canonical_sh = canonical.book.sheet_by_index(canonical.sheet_idx)
    if (canonical_sh.rowinfo_map or {}).get(0):
        ri = canonical_sh.rowinfo_map[0]
        if ri.height:
            ws.row_dimensions[1].height = ri.height / 20.0
    ws.freeze_panes = "A2"

    # ---- 数据行 ----
    out_row = 2
    n_rows_written = 0
    n_images_written = 0
    n_dedup_skipped = 0

    # 全局已见快递单号 (每个汇总文件 / 月份独立)
    seen_tracking: set = set()

    # 为每个 (file, sheet) 跟踪 src_row -> out_row 映射, 用于翻译合并单元格
    # key: (id(pf)); value: dict src_row_0based -> out_row
    src_to_out_per_pf: list[tuple[ParsedFile, dict]] = []

    for pf in parsed_files:
        src_to_out: dict = {}
        multi_sheet = file_sheet_count[pf.src_file] > 1
        src_label = (
            f"{pf.src_file} / {pf.src_sheet}" if multi_sheet else pf.src_file
        )
        for prow in pf.rows:
            try:
                # ---- 去重: 按快递单号 ----
                if pf.tracking_col >= 0 and pf.tracking_col < len(prow.cells):
                    tn_raw = prow.cells[pf.tracking_col][0]
                    tn_norm = _normalize_tracking(tn_raw)
                    if tn_norm is not None:
                        if tn_norm in seen_tracking:
                            logger.info(
                                "[%s] 第 %d 行 快递单号=%s 已存在, 跳过",
                                src_label, prow.src_row_1based, tn_raw,
                            )
                            n_dedup_skipped += 1
                            continue
                        seen_tracking.add(tn_norm)

                src_to_out[prow.src_row_0based] = out_row
                cells = list(prow.cells)
                if len(cells) > n_canon_cols:
                    cells = cells[:n_canon_cols]
                while len(cells) < n_canon_cols:
                    cells.append((None, 0))

                for col_idx, (val, xfi) in enumerate(cells, start=1):
                    is_image_col = (col_idx - 1) == canonical_image_col
                    cell = ws.cell(
                        row=out_row,
                        column=col_idx,
                        value=None if is_image_col else val,
                    )
                    font, align, border, nfmt = style_cache.get(pf.book, xfi)
                    _apply_style(cell, font, align, border, nfmt)

                ws.cell(row=out_row, column=n_canon_cols + 1, value=src_label)
                ws.cell(row=out_row, column=n_canon_cols + 2, value=prow.src_row_1based)

                if prow.image is not None and canonical_image_col >= 0:
                    fmt, image_bytes = prow.image
                    buf, _ext = _prepare_image_bytes(image_bytes, fmt)
                    try:
                        xl_img = XLImage(buf)
                        xl_img.width = _IMG_BOX_PX
                        xl_img.height = _IMG_BOX_PX
                        cell_addr = f"{get_column_letter(canonical_image_col + 1)}{out_row}"
                        ws.add_image(xl_img, cell_addr)
                        n_images_written += 1
                    except Exception as e:
                        logger.error(
                            "[%s] 第 %d 行图片嵌入失败 (%s, %d 字节): %s",
                            src_label, prow.src_row_1based, fmt, len(image_bytes), e,
                        )
                        logger.debug(traceback.format_exc())

                row_height = prow.row_height_pt
                if prow.image is not None:
                    row_height = max(row_height or 0, _IMG_ROW_HEIGHT_PT)
                if row_height:
                    ws.row_dimensions[out_row].height = row_height

                out_row += 1
                n_rows_written += 1
            except Exception as e:
                logger.error(
                    "[%s] 写入第 %d 行 (源 %d 行) 失败: %s",
                    src_label, out_row, prow.src_row_1based, e,
                )
                logger.debug(traceback.format_exc())
        src_to_out_per_pf.append((pf, src_to_out))

    # ---- 合并单元格 (按 src_row -> out_row 翻译) ----
    n_merges_applied = 0
    n_merges_skipped = 0
    for pf, src_to_out in src_to_out_per_pf:
        for rlo, rhi, clo, chi in pf.merges:
            # 跳过表头里 (rlo == 0) 的合并: 我们的输出表头只占 1 行
            if rlo == 0 and rhi == 1:
                # 单行合并: 如果跨多列, 应用到输出表头
                end_col = min(chi, n_canon_cols + 1)
                if end_col - clo > 1:
                    try:
                        ws.merge_cells(
                            start_row=1, start_column=clo + 1,
                            end_row=1, end_column=end_col,
                        )
                        n_merges_applied += 1
                    except Exception as e:
                        logger.warning(
                            "[%s] 表头合并 col=[%d,%d) 失败: %s",
                            pf.src_file, clo, chi, e,
                        )
                        n_merges_skipped += 1
                continue
            # 数据行合并: 检查所有 src 行都被导出且连续
            src_rows = list(range(rlo, rhi))
            out_rows = [src_to_out.get(r) for r in src_rows]
            if any(o is None for o in out_rows):
                # 有行被空跳过, 无法合并
                n_merges_skipped += 1
                logger.debug(
                    "[%s] 合并 (rows=%d-%d cols=%d-%d) 跳过: 部分源行未导出",
                    pf.src_file, rlo + 1, rhi, clo + 1, chi,
                )
                continue
            if any(out_rows[i] - out_rows[i - 1] != 1 for i in range(1, len(out_rows))):
                n_merges_skipped += 1
                logger.warning(
                    "[%s] 合并 (rows=%d-%d cols=%d-%d) 跳过: 源行映射到输出后不连续",
                    pf.src_file, rlo + 1, rhi, clo + 1, chi,
                )
                continue
            end_col = min(chi, n_canon_cols + 1)  # 不越过 n_canon_cols
            if end_col <= clo:
                continue
            try:
                ws.merge_cells(
                    start_row=out_rows[0], start_column=clo + 1,
                    end_row=out_rows[-1], end_column=end_col,
                )
                n_merges_applied += 1
            except Exception as e:
                logger.warning(
                    "[%s] 合并 (out rows=%d-%d cols=%d-%d) 失败: %s",
                    pf.src_file, out_rows[0], out_rows[-1], clo + 1, end_col, e,
                )
                n_merges_skipped += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info(
        "已写入 %s: 数据行=%d, 嵌入图片=%d, 去重跳过=%d, "
        "合并单元格已应用=%d, 跳过=%d, 唯一快递单号=%d",
        out_path, n_rows_written, n_images_written, n_dedup_skipped,
        n_merges_applied, n_merges_skipped, len(seen_tracking),
    )


# ====== 月份分组 + 主流程 ===================================================

# 例: "3.10做单表格-1杨(1).xls" -> month=3, day=10
_FNAME_RE = re.compile(r"^\s*(\d{1,2})\s*[\.\u3002]\s*(\d{1,2})(?=\D|$)")


def group_files_by_month(src_dir: Path, logger: logging.Logger) -> dict[int, list[Path]]:
    groups: dict[int, list[tuple[int, str, Path]]] = defaultdict(list)
    skipped: list[str] = []
    for p in sorted(src_dir.iterdir()):
        if not p.is_file():
            continue
        if p.name.startswith(".") or p.name.startswith("~$"):
            continue
        if p.suffix.lower() != ".xls":
            logger.info("跳过非 .xls 文件: %s", p.name)
            continue
        m = _FNAME_RE.match(p.name)
        if not m:
            skipped.append(p.name)
            continue
        month = int(m.group(1))
        day = int(m.group(2))
        groups[month].append((day, p.name, p))

    if skipped:
        logger.warning("以下 %d 个文件无法识别月份, 已跳过:", len(skipped))
        for s in skipped:
            logger.warning("  - %s", s)

    final: dict[int, list[Path]] = {}
    for m, items in groups.items():
        items.sort(key=lambda x: (x[0], x[1]))
        final[m] = [it[2] for it in items]
    return final


def main():
    ap = argparse.ArgumentParser(description="按月合并 .xls (含图片+格式+合并) 为 .xlsx")
    ap.add_argument(
        "src", nargs="?",
        default="/Users/lishengsheng/Documents/铝合金代发",
        help="源目录",
    )
    ap.add_argument(
        "out", nargs="?", default=None,
        help="输出目录 (默认: 源目录/_合并)",
    )
    args = ap.parse_args()

    src_dir = Path(args.src).expanduser().resolve()
    out_dir = Path(args.out).expanduser().resolve() if args.out else src_dir / "_合并"
    out_dir.mkdir(parents=True, exist_ok=True)

    log_path = out_dir / "merge.log"
    logger = setup_logger(log_path)

    logger.info("=" * 70)
    logger.info("源目录  : %s", src_dir)
    logger.info("输出目录: %s", out_dir)
    logger.info("日志文件: %s", log_path)
    logger.info("=" * 70)

    if not src_dir.is_dir():
        logger.error("源目录不存在: %s", src_dir)
        sys.exit(1)

    groups = group_files_by_month(src_dir, logger)
    if not groups:
        logger.warning("没找到任何可处理的 .xls 文件")
        return

    for month in sorted(groups.keys()):
        files = groups[month]
        logger.info("")
        logger.info("---------- 处理 %d 月 (共 %d 个文件) ----------", month, len(files))
        parsed_list: list[ParsedFile] = []
        n_files_used = 0
        for f in files:
            logger.info("解析: %s", f.name)
            sheets = parse_xls_file(f, logger)
            if sheets:
                parsed_list.extend(sheets)
                n_files_used += 1
        if not parsed_list:
            logger.warning("%d 月没有可用数据", month)
            continue
        out_file = out_dir / f"{month}月汇总.xlsx"
        total_rows = sum(len(pf.rows) for pf in parsed_list)
        logger.info(
            "%d 月: %d 个文件 (%d 个 sheet), %d 条数据行 (去重前), 写入 %s",
            month, n_files_used, len(parsed_list), total_rows, out_file.name,
        )
        write_merged_xlsx(out_file, parsed_list, logger)

    logger.info("")
    logger.info("全部完成。详细日志: %s", log_path)


if __name__ == "__main__":
    main()
